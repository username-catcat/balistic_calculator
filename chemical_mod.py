import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import math
import pandas as pd
from openpyxl import load_workbook
#Данные для массива кинематической вязкости(500 метров= 1 позиция)
class output:
    @staticmethod
    def every_value(tirray, aarray, verray, herray):
        plt.plot(tirray, verray, label="velocity")
        plt.plot(tirray, herray, label="height")
        plt.plot(tirray, aarray, label="acelleration")
        plt.xlabel("total time")
        plt.ylabel("value")
        plt.title("dependence by time")
        plt.legend()
        plt.show()

    @staticmethod
    def trace(derray, herray):
        plt.plot(derray, herray)
        plt.ylabel("height")
        plt.xlabel("distance")
        plt.title("trace")
        plt.show()

    @staticmethod
    def acelleration(tirray, aarray):
        plt.plot(tirray, aarray)
        plt.ylabel("acelleration")
        plt.xlabel("total time")
        plt.title("acelleration by time")
        plt.show()

    @staticmethod
    def velocity(tirray, verray):
        plt.plot(tirray, verray)
        plt.ylabel("velocity")
        plt.xlabel("total time")
        plt.title("velocity by time")
        plt.show()

    @staticmethod
    def height(tirray, herray):
        plt.plot(tirray, herray)
        plt.ylabel("height")
        plt.xlabel("total time")
        plt.title("height by time")
        plt.show()
file =("C:\database\database.xlsx")
wb = load_workbook(file)
ws = wb.active
oxidizer=""
lfuel=""
print("would you like to use chemisty module?")
chemistry = int(input())
print("insert rocket height (m)")
h = float(input())
print("insert rocket mass (kg)")
mass = float(input())
print("insert fuel mass (kg)")
fuel = float(input())
num=2
formula=""
if chemistry == True :
    print("select composition state")
    state = input()
    print("\navailable compositions:")
    if (state == "solid"):
        while(ws.cell(row=num, column=1).value!= 0):
            if (ws.cell(row=num, column=9).value == state): print(ws.cell(row=num, column=1).value)
            num+=1
        num=2
        print("\nselect composition by name")
        name=input()
        while(ws.cell(row=num, column=1).value!=name):num+=1
    else:
        while(ws.cell(row=num, column=1).value!= 0):
            if (ws.cell(row=num, column=9).value == state): 
                print(ws.cell(row=num, column=2).value, ws.cell(row=num, column=3).value)
            num+=1
        num=2
        print("\nselect composition by formula")
        formula=input()
        formula=formula.split()
        lfuel= formula[1]
        oxidizer=formula[0]
        while(ws.cell(row=num, column=1).value!=0):
            if (ws.cell(row=num, column=2).value!=oxidizer):
                if(ws.cell(row=num, column=3).value!=lfuel):break
            num+=1
    specific_impulse = float(ws.cell(row=num, column=8).value)
else :
    print("insert specific impulse (N*s)")
    specific_impulse = float(input())
#temporary solution, while mass flow rate is not available
print("insert fuel burning time (s)")
ALT = float(input())
print("insert angle on start")
angle = float(input())
angle = math.radians(angle)
g = 9.81
horizontalvelocity = 0.0
verticalvelocity = 0.0
height = 0.0
dtime = 0.001
time = 0.0
#Данные для массива кинематической вязкости(500 метров= 1 позиция)
kin= [
    1.46*(10**-5),
    1.52*(10**-5),
    1.58*(10**-5),
    1.65*(10**-5),
    1.71*(10**-5),
    1.79*(10**-5)
    ]
verticalacelleration = 0.0
horizontalacelleration = 0.0
distance = 0.0
power = specific_impulse / ALT
mass = mass + fuel
dfuel = fuel / ALT * dtime
maxheight = height
maxvelocity = 0.0
maxacelleration = 0.0
value = []
herray = []
tirray = []
derray = []
aarray = []
verray = []
iteration = 0
Cx = 0.25
Recrit=51*(10**-5*h)**-1.039
r = 0.03
airdensity = 1.3
volume = math.pi * (r**2) * h
Resistance = 0
he=0
di=0
Cf=0
while height >= 0:
    maxvelocity = max(maxvelocity, math.sqrt((horizontalvelocity**2) + (verticalvelocity**2)))
    maxacelleration = max(maxacelleration, math.sqrt((horizontalacelleration**2) + (verticalacelleration**2)))
    maxheight = max(maxheight, height)
    if height>500:
        he=1
        if height>1000:
            he=2
            if height>1500:
                he=3
                if height>2000:
                    he=4
                    if height>2500:
                        he=5
    if iteration > 1:
        verticalacelleration = (power * math.cos(angle)- (Resistance* (verticalvelocity/ math.sqrt((horizontalvelocity**2) + (verticalvelocity**2))))/ mass) - g
        horizontalacelleration = (power * math.sin(angle)- (Resistance* (horizontalvelocity/ math.sqrt((horizontalvelocity**2) + (verticalvelocity**2))))/ mass)
        Resistance = Cx * ((volume ** (2 / 3))* (airdensity * ((horizontalvelocity**2) + (verticalvelocity**2)) / 2)) + Cf * ((volume ** (2 / 3))* (airdensity * ((horizontalvelocity**2) + (verticalvelocity**2)) / 2))
    else:
        verticalacelleration = power * math.cos(angle) / mass - g
        horizontalacelleration = power * math.sin(angle) / mass
    horizontalvelocity = horizontalvelocity + (horizontalacelleration * dtime)
    verticalvelocity = verticalvelocity + (verticalacelleration * dtime)
    height = height + (verticalvelocity * dtime)
    distance = distance + (horizontalvelocity * dtime)
    time += dtime
    herray.append(height)
    derray.append(distance)
    aarray.append(verticalacelleration)
    verray.append(verticalvelocity)
    tirray.append(time)
    mass = mass - dfuel
    while(di<h):
        LRc=di*math.sqrt((horizontalvelocity**2) + (verticalvelocity**2))/kin[he]
        if LRc>Recrit: Cf= 0.032*(10**-5*h)**0.2
        if LRc<10000: Cf= 0.0148
        di+=0.0001
    if time >= ALT and power != 0:
        power = 0
        dfuel = 0
    iteration += 1
print(
    f"max velocity"
    + str(maxvelocity)
    + "    max height "
    + str(maxheight)
    + "    max acelleration "
    + str(maxacelleration)
    + "    flight time "
    + str(time)
    + "    distance "
    + str(distance)
)
output.every_value(tirray, aarray, verray, herray)
output.trace(derray, herray)
